from tkinter import *
from tkinter import messagebox
import ast
import openpyxl
from openpyxl import load_workbook

window = Tk()
window.title("SignUp")
window.geometry("925x500+300+200")
window.config(bg="#fff")
window.resizable(False,False)

def signup():
    username= user.get()
    password = code.get()
    confirm_password = confirm_code.get()

    if password == confirm_password:
        try:
            wb = openpyxl.load_workbook("login.xlsx")
            sheet = wb.active
            sheet.append([username,password])
            wb.save("login.xlsx")
            

        except:
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet["A1"] = "Username"
            sheet["B1"] = "Password"

            sheet.append([username,password])
            wb.save("login.xlsx")
        messagebox.showinfo("successfully","Resistered")

    else:
        messagebox.showerror("Error:","username and passsword should be same")
    


img = PhotoImage(file="login.png")
Label(window,border=0,image=img,bg="white",fg="white").place(x=50,y=50)


frame = Frame(window,width=350,height=390,bg = "#fff")
frame.place(x=480,y=50)


heading = Label(frame,text="Sign Up",fg="blue",bg = "white",font=('Microsoft yahei UI light',23,"bold"))
heading.place(x= 100,y=5)

###-----------------------------------------------
def on_enter(e):
    user.delete(0,'end')
    
def on_leave(e):
    if user.get() =="":
        user.insert(0,'Username')

user = Entry(frame,width=25,fg='black',border=0,bg= "white",font=('Microsoft yahei UI light',11))
user.place(x=30,y=80)
user.insert(0,'username')
user.bind("<FocusIn>",on_enter)
user.bind("<FocusOut>",on_leave)


###-----------------------------------------------
def on_enter(e):
    code.delete(0,'end')
    
def on_leave(e):
    if code.get() =="":
        code.insert(0,'Password')

code = Entry(frame,width=25,fg='black',border=0,bg= "white",font=('Microsoft yahei UI light',11))
code.place(x=30,y=150)
code.insert(0,'Password')
code.bind("<FocusIn>",on_enter)
code.bind("<FocusOut>",on_leave)



###-----------------------------------------------
def on_enter(e):
    confirm_code.delete(0,'end')
    
def on_leave(e):
    if confirm_code.get() =="":
        confirm_code.insert(0,'Confirm Password')

confirm_code = Entry(frame,width=25,fg='black',border=0,bg= "white",font=('Microsoft yahei UI light',11))
confirm_code.place(x=30,y=220)
confirm_code.insert(0,'Confirm Password')
confirm_code.bind("<FocusIn>",on_enter)
confirm_code.bind("<FocusOut>",on_leave)


#----------------------------------------------------------------------------
Button(frame,width=39,pady=7,command=signup,text="Sing Up",bg='lightblue',fg="white",border=0).place(x = 35,y = 280)

label = Label(frame,text="I have an account",fg='black',bg="white",font=('Microsoft yahei UI light',11))
label.place(x = 75,y=335)

signin= Button(frame,width=7,text='sign in',border=0,bg = 'white',cursor='hand2',fg="#57a1f8")
signin.place(x =200,y = 340)


window.mainloop()