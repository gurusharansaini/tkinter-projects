from openpyxl import load_workbook
import tkinter  as tk
import openpyxl

headers = ["Name","Course","Semester","Form No","contact","Email id","Address"]



def check():
    try:
        wb = load_workbook('excel.xlsx')
    except:
        wb = openpyxl.Workbook()
        sheet = wb.active
        for i in range(0,7):
            sheet[f"{chr(65+i)}1"] = headers[i]
        wb.save("excel.xlsx")
        
        

def submit():
    check()
    wb = load_workbook("excel.xlsx")
    sheet = wb.active
    lst = [name.get(),course.get(),semester.get(),Form.get(),Contact.get(),email.get(),address.get()]
    sheet.append(lst)
    wb.save('excel.xlsx')

    name.set("")
    course.set("")
    semester.set("")
    Form.set("")
    Contact.set("")
    email.set("")
    address.set("")
    
    
    
    


    
     



if __name__=="__main__":
    check()

    root = tk.Tk()
    root.geometry("500x350")
    root.title("Resistration Form")
    root.configure(bg='lightgreen')
    name = tk.StringVar()
    course = tk.StringVar()
    semester = tk.StringVar()
    Form = tk.StringVar()
    Contact = tk.StringVar()
    email = tk.StringVar()
    address = tk.StringVar()



    lbl_student_Name = tk.Label(root,text="Name",bg = "lightgreen",font=("goudy old style",20))
    lbl_student_Name.grid(row=0,column=0)

    lbl_student_course = tk.Label(root,text="Course",bg = "lightgreen",font=("goudy old style",20))
    lbl_student_course.grid(row=1,column=0)

    lbl_student_sem = tk.Label(root,text="Semester",bg = "lightgreen",font=("goudy old style",20))
    lbl_student_sem.grid(row=2,column=0)

    lbl_student_formNo = tk.Label(root,text="Form No.",bg = "lightgreen",font=("goudy old style",20))
    lbl_student_formNo.grid(row=3,column=0)

    lbl_student_Contact = tk.Label(root,text="Contact No.",bg = "lightgreen",font=("goudy old style",20))
    lbl_student_Contact.grid(row=4,column=0)

    lbl_student_emai = tk.Label(root,text="Email id",bg = "lightgreen",font=("goudy old style",20))
    lbl_student_emai.grid(row=5,column=0)

    lbl_student_addr = tk.Label(root,text="Address",bg = "lightgreen",font=("goudy old style",20))
    lbl_student_addr.grid(row=6,column=0)



    entry_student_Name = tk.Entry(root,textvariable=name,font=("goudy old style",25))
    entry_student_Name.grid(row=0,column=1)

    entry_student_course = tk.Entry(root,textvariable=course,font=("goudy old style",25))
    entry_student_course.grid(row=1,column=1)

    entry_student_sem = tk.Entry(root,textvariable=semester,font=("goudy old style",25))
    entry_student_sem.grid(row=2,column=1)

    entry_student_formNo = tk.Entry(root,textvariable=Form,font=("goudy old style",25))
    entry_student_formNo.grid(row=3,column=1)

    entry_student_Contact = tk.Entry(root,textvariable=Contact,font=("goudy old style",25))
    entry_student_Contact.grid(row=4,column=1)

    entry_student_emai = tk.Entry(root,textvariable=email,font=("goudy old style",25))
    entry_student_emai.grid(row=5,column=1)

    entry_student_addr = tk.Entry(root,textvariable=address,font=("goudy old style",25))
    entry_student_addr.grid(row=6,column=1)

    btn_submit = tk.Button(root,text="Submit",command=submit,fg="black",bg="white",width=20)
    btn_submit.grid(row=7,column=1)
   

    root.mainloop()